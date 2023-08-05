from typing import Union, Optional, List, Dict

from divinegift import main

import os
from datetime import datetime, timedelta, date
try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    raise ImportError("openpyxl isn't installed. Run: pip install -U openpyxl")
try:
    import xlrd
except ImportError:
    raise ImportError("xlrd isn't installed. Run: pip install -U xlrd")
try:
    import xlsxwriter
except ImportError:
    raise ImportError("xlsxwriter isn't installed. Run: pip install -U xlsxwriter")
try:
    import xlwt
except ImportError:
    raise ImportError("xlwt isn't installed. Run: pip install -U xlwt")


def create_excel(list_: List[Union[Dict, List]], fname: str, header: Optional[List] = None,
                 column_width: Optional[List] = None, worksheet: str = 'Sheet', start_row: int = 0, borders=False):
    """
    Создает Excel-файл в формате xlsx
    :param list_: Входной список
    :param fname: Имя файла с путем
    :param column_width: Список с шириной столбцов
    :param worksheet: Имя вкладки
    :param start_row: С какой строки начинать писать
    :return:
    """
    # Если указанной папки не существует - создаем ее (включая родительские, если нужно)
    fd, fn = os.path.split(fname)
    if fd:
        main.check_folder_exist(fd)
    # Создаем книгу
    if fname.split('.')[-1] == 'xlsx':
        create_xlsx(fname, list_, column_width, header, worksheet, start_row, borders)
    else:
        create_xls(fname, list_, column_width, header, worksheet, start_row, borders)


def create_xls(filename: str, data: List[Union[Dict, List]], column_width: Optional[List], header: Optional[List],
               sheet_name: str = 'Sheet', start_row: int = 0, add_borders: bool = False):
    if not filename.endswith('xls'):
        filename += '.xls'
    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheet_name)

    common_style = xlwt.XFStyle()
    font = xlwt.Font()
    common_style.alignment.wrap = 1
    if add_borders:
        common_borders = xlwt.Borders()
        common_borders.bottom = xlwt.Borders.THIN
        common_borders.top = xlwt.Borders.THIN
        common_borders.left = xlwt.Borders.THIN
        common_borders.right = xlwt.Borders.THIN

        common_style.borders = common_borders
    common_style.font = font

    if header:
        # Устанавливаем стиль для заголовков
        header_style = xlwt.XFStyle()
        header_style.alignment.wrap = 1
        header_font = xlwt.Font()
        header_font.bold = True

        header_borders = xlwt.Borders()
        if add_borders:
            header_borders.top = xlwt.Borders.THIN
            header_borders.left = xlwt.Borders.THIN
            header_borders.right = xlwt.Borders.THIN
        header_borders.bottom = xlwt.Borders.THIN
        header_style.borders = header_borders
        header_style.font = header_font

        # Заполняем шапку
        for c_index, c_value in enumerate(header):
            ws.write(0, c_index, c_value, header_style)
        start_row += 1

    if column_width:
        # Устанавливаем ширину столбцов
        for c_index, c_width in enumerate(column_width):
            ws.col(c_index).width = 256 * c_width

    # Вставляем значения
    for r_index, row in enumerate(data):
        try:
            columns = list(row.values())
        except AttributeError:
            columns = row
        for c_index, c_value in enumerate(columns):
            # Если значение типа datetime - преобразовываем в строку вида dd.MM.yyyy HH:mm:ss
            if isinstance(c_value, datetime):
                common_style.num_format_str = 'dd.mm.yyyy hh:mm:ss'
            # Если значение типа date - преобразовываем в строку вида dd.MM.yyyy
            elif isinstance(c_value, date):
                common_style.num_format_str = 'dd.mm.yyyy'
            elif isinstance(c_value, int):
                common_style.num_format_str = 'General'
            elif isinstance(c_value, float):
                common_style.num_format_str = 'General'
            else:
                common_style.num_format_str = 'General'

            ws.write(start_row + r_index, c_index, c_value, common_style)
    wb.save(filename)


def create_xlsx(filename: str, data: List[Union[Dict, List]], column_width: Optional[List], header: Optional[List],
               sheet_name: str = 'Sheet', start_row: int = 0, add_borders: bool = False):
    if not filename.endswith('xlsx'):
        filename += '.xlsx'
    wb = xlsxwriter.Workbook(filename)
    ws = wb.add_worksheet(sheet_name)

    common_style = wb.add_format({'text_wrap': True})
    if add_borders:
        common_style.set_border()

    if header:
        header_style = wb.add_format({'bold': 1, 'text_wrap': True})
        if add_borders:
            header_style.set_border()
        else:
            header_style.set_bottom()
        # Заполняем шапку
        for c_index, c_value in enumerate(header):
            ws.write(0, c_index, c_value, header_style)
        start_row += 1

    if column_width:
        for c_index, c_width in enumerate(column_width):
            ws.set_column(c_index, c_index, c_width)

    # Вставляем значения
    for r_index, row in enumerate(data):
        try:
            columns = list(row.values())
        except AttributeError:
            columns = row
        for c_index, c_value in enumerate(columns):
            # Если значение типа datetime - преобразовываем в строку вида dd.MM.yyyy HH:mm:ss
            if isinstance(c_value, datetime):
                common_style.num_format = 'dd.mm.yyyy hh:mm:ss'
            # Если значение типа date - преобразовываем в строку вида dd.MM.yyyy
            elif isinstance(c_value, date):
                common_style.num_format = 'dd.mm.yyyy'
            elif isinstance(c_value, int):
                common_style.num_format = 'General'
            elif isinstance(c_value, float):
                common_style.num_format = 'General'
            else:
                common_style.num_format = 'General'

            ws.write(r_index + start_row, c_index, c_value, common_style)
    wb.close()


def read_excel(filename: str, excel_header: Union[list, dict], sheet_name: str = None,
               int_columns: list = None, date_columns: list = None, start_row: int = 1):
    # Считываем Excel-файл
    excel_arr = []
    if isinstance(excel_header, list):
        columns = excel_header
    else:
        columns = [x[1] for x in list(excel_header.values())]

    if filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name] if sheet_name else wb.active
        records = [[x.value for x in row] for row in ws.iter_rows(min_row=start_row)]
    else:
        wb = xlrd.open_workbook(filename)
        ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
        records = [ws.row_values(i) for i in range(start_row - 1, ws.nrows)]

    int_columns = int_columns if int_columns else []
    date_columns = date_columns if date_columns else []

    # По каждой строчке
    for i, r in enumerate(records):
        # Сцепляем заголовки (английские) из настроек с данными из Excel-файла
        tmp_row = dict(zip(columns, r))

        # Костыль от преобразований данных экселем после ручных правок
        for c in columns:
            # Костыль для столбцов, которые должны быть int
            if tmp_row.get(c) is not None:
                if c in int_columns and not isinstance(tmp_row.get(c), int):
                    tmp_row[c] = int(tmp_row[c])
                # Костыль для столбцов, которые должны быть datetime
                if c in date_columns and not isinstance(tmp_row.get(c), datetime):
                    try:
                        tmp_row[c] = main.parse_date(tmp_row[c])
                    except TypeError:
                        # Костыль, если дата в экселе не текст, а число
                        tmp_row[c] = datetime(1899, 12, 31) + timedelta(days=tmp_row[c] - 1)
                    except ValueError:
                        tmp_row[c] = None
                # Костыль для столбцов, которые должны быть str
                if c not in date_columns and c not in int_columns and not isinstance(tmp_row.get(c), str):
                    tmp_row[c] = str(tmp_row.get(c))
        excel_arr.append(tmp_row)

    return excel_arr    


if __name__ == '__main__':
    pass
