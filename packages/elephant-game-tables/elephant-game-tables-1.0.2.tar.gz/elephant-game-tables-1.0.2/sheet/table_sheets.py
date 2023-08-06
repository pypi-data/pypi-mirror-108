from openpyxl import load_workbook

from sheet.step.find_col import FindTableColStep
from sheet.step.find_row import FindTableRowStep
from sheet.step.folding import FoldingStep
from sheet.step.merge_reference_sheet import MergeReferenceSheetStep
from sheet.step.parse_data import ParseDataStep
from sheet.step.parse_field import ParseFieldStep
from sheet.step.parse_phase import ParsePhase
from sheet.step.parse_step import TableSheetParser
from sheet.step.validate import TableSheetValidateStep
from util.type_utils import TypeUtils


class TableSheet:
    @classmethod
    def load_file(cls, table_sheet_class, manager, file_path):
        table_sheets = {}
        work_book = load_workbook(filename=file_path)
        for sheet_name in work_book.sheetnames:
            new_table_sheet = table_sheet_class(manager)
            if new_table_sheet.initialize(file_path, sheet_name, work_book[sheet_name]):
                table_sheets[sheet_name] = new_table_sheet
        return table_sheets

    def __init__(self, sheet_manager):
        self.sheet_manager = sheet_manager

        self.sheet_file = ""
        self.sheet_name = ""
        self.work_sheet = None

        self.contents = None
        self.fields = []
        self.referencing_sheets = []

        self.parser = TableSheetParser(self)
        self._init_parser()

    def initialize(self, file_path, sheet_name, work_sheet):
        self.sheet_file = file_path
        self.sheet_name = sheet_name
        self.work_sheet = work_sheet
        return self.parse(ParsePhase.LoadTable)

    def append_parse_step(self, parse_step):
        self.parser.append_step(parse_step)

    def _init_parser(self):
        self.append_parse_step(TableSheetValidateStep())
        self.append_parse_step(FindTableRowStep())
        self.append_parse_step(FindTableColStep())
        self.append_parse_step(ParseFieldStep())
        self.append_parse_step(ParseDataStep())
        self.append_parse_step(FoldingStep())
        self.append_parse_step(MergeReferenceSheetStep())

    def parse(self, parse_phase):
        return self.parser.start_parse(parse_phase)

    def get_contents(self):
        return self.contents

    def get_fields(self):
        fields_ret = []
        for field in self.fields:
            if not self._filter_field(field):
                continue
            fields_ret.append({
                "name": field.name,
                "type": TypeUtils.get_typescript_type_of(field.type)
            })
        return fields_ret

    @staticmethod
    def _filter_field(field):
        if field.name == "":
            return False
        if field.name == "#":
            return False
        if field.name.find("__") != -1:
            return False
        if field.type == "":
            return False
        if field.type == "#":
            return False
        return True
