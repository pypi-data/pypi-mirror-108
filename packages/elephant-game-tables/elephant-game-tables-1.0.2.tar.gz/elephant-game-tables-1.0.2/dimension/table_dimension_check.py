import os
from datetime import datetime

from openpyxl import load_workbook

from transform.resource_manager import GameResourceManager
from util.print_utils import print_block
from util.table_utils import get_sheet_dimension, get_sheet_name, is_file_filtered


class TableDimensionCheck:
    def __init__(self, project_dir):
        self.project_dir = GameResourceManager(project_dir)

    @staticmethod
    def _get_all_table_files(table_dir):
        table_files = []
        for file in os.listdir(table_dir):
            file_path = os.path.join(table_dir, file)
            if not is_file_filtered(file_path):
                table_files.append(file_path)
        return table_files

    def check_all(self, table_dir):
        print_block("check all files : begin!")
        all_table_files = self._get_all_table_files(table_dir)
        for table_file in all_table_files:
            if not self.check_table_file(table_file):
                break
        print_block("check all files : end!")

    def check_table_file(self, file_path):
        file_name = os.path.basename(file_path)
        wb = load_workbook(file_path)
        file_start_time = datetime.now()
        for sheet in wb.worksheets:
            sheet_start_time = datetime.now()
            dimension_sheet_file_path, dimension_sheet = self.get_dimension_sheet(file_name, sheet.title)
            error_detail = self.check_sheet_content(dimension_sheet, sheet)
            time_used = datetime.now() - sheet_start_time
            if error_detail:
                print_block("check table sheet fail!", {
                    "sheetName": sheet.title,
                    "srcFile": file_path,
                    "dstFile": dimension_sheet_file_path,
                    "error": error_detail,
                    "timeUsed": time_used
                })
                return False
            print_block("check table sheet success!", {
                "sheetName": sheet.title,
                "srcFile": file_path,
                "dstFile": dimension_sheet_file_path,
                "timeUsed": time_used
            })

        print_block("check table file success!", {
            "srcFile": file_path,
            "timeUsed": datetime.now() - file_start_time
        })
        return True

    def get_dimension_sheet(self, file_name, sheet_name):
        dimension = get_sheet_dimension(sheet_name)
        sheet_name_org = get_sheet_name(sheet_name)
        dimension_table_file_path = os.path.join(self.project_dir.get_table_dir(), os.path.join(dimension, file_name) if dimension else file_name)
        wb = load_workbook(dimension_table_file_path)
        return dimension_table_file_path, wb.get_sheet_by_name(sheet_name_org)

    def check_sheet_content(self, src_sheet, dst_sheet):
        for row_index in range(1, src_sheet.max_row):
            for col_index in range(1, src_sheet.max_column):
                x = src_sheet[row_index][col_index]
                y = dst_sheet[row_index][col_index]
                if x.value != y.value:
                    return {
                        "row": {
                            "index": row_index,
                            "value": x.value
                        },
                        "col": {
                            "index": col_index,
                            "value": y.value
                        },
                    }
        return None
