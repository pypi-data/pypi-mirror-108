import os
import shutil

from openpyxl import load_workbook

from util.table_utils import get_sheet_dimension, get_sheet_name, is_file_filtered


class TableSpreader:
    def spread_all(self, target_dir):
        for file in os.listdir(target_dir):
            table_file_path = os.path.join(target_dir, file)
            if not os.path.isfile(table_file_path) or is_file_filtered(table_file_path):
                continue
            self._spread_table_sheet(table_file_path)

    @staticmethod
    def _get_dimensions(table_file_path):
        wb = load_workbook(table_file_path)
        dimensions = set()
        for sheet in wb.worksheets:
            dimension_name = get_sheet_dimension(sheet.title)
            if dimension_name:
                dimensions.add(dimension_name)
        return dimensions

    @staticmethod
    def _normalize_table(dimension, table_file_path):
        wb = load_workbook(table_file_path)
        for sheet in wb.worksheets:
            if dimension not in sheet.title:
                wb.remove(sheet)

        for sheet in wb.worksheets:
            sheet.title = get_sheet_name(sheet.title)

        wb.save(table_file_path)

    @staticmethod
    def _normalize_table_default(table_file_path):
        wb = load_workbook(table_file_path)
        for sheet in wb.worksheets:
            if "@" in sheet.title:
                wb.remove(sheet)

        wb.save(table_file_path)

    def _spread_table_sheet(self, table_file_path):
        dimensions = self._get_dimensions(table_file_path)
        file_name = os.path.basename(table_file_path)
        process_map = {}
        for dimension in dimensions:
            dimension_path = os.path.join(os.path.dirname(table_file_path), dimension.lower())
            if not os.path.exists(dimension_path):
                os.mkdir(dimension_path)
            dimension_file_path = os.path.join(dimension_path, file_name)
            process_map[dimension] = dimension_file_path
            shutil.copy(table_file_path, dimension_file_path)
            print(f"spread table file from [{table_file_path}] to [{dimension_file_path}]")

        if len(process_map) > 0:
            for dimension, dimension_table_file_path in process_map.items():
                self._normalize_table(dimension, dimension_table_file_path)
                print(f'normalize table of dimension [{dimension}] with file => [{dimension_table_file_path}]')
            self._normalize_table_default(table_file_path)
            print(f'normalize table of default with file => [{table_file_path}]')
