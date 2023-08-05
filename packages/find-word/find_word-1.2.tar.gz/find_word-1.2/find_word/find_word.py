import os
import shutil
import subprocess
import docx
import openpyxl
import PyPDF2 
import warnings

class search_word:

    #lists for collecting information
    find_file = []
    spisok_file = []
    exc = []

    #the main function for finding words in files
    def search(self, disk, format_files, word):           
        #txt
        if format_files == 'txt' or format_files == 'all': 
            for adress, dirs, files in os.walk(disk + ':\\'):   
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith('.txt') and '$' not in s:
                        self.spisok_file.append(s)
            for x in self.spisok_file: 
                with open(x) as r:
                    try:
                        for line in r:
                            if word in line:   
                                self.find_file.append(x)  
                                break
                    except Exception as fail:
                        self.exc.append(r) 

        #docx
        if format_files == 'doc' or format_files == 'all': 
            for adress, dirs, files in os.walk(disk + ':\\'):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith ('.docx') or file.endswith('.doc') and '$' not in s:
                        self.spisok_file.append(s)     
            for x in self.spisok_file: 
                with open(x) as r:
                    try:
                        doc = docx.Document(x)
                        text = []
                        for paragraph in doc.paragraphs:
                            text.append(paragraph.text)    
                        for line in text:
                            if word in line:                        
                                self.find_file.append(x)
                                break
                    except Exception as fail:
                        self.exc.append(r)   

        #exel
        if format_files == 'exl' or format_files == 'all': 
            for adress, dirs, files in os.walk(disk + ':\\'):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith('.xls') or file.endswith('.xlsx') and '$' not in s:
                        self.spisok_file.append(s)
            for x in self.spisok_file: 
                with open(x) as r:
                    try:
                        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                        path = x
                        wb_obj = openpyxl.load_workbook(path)
                        sheet_obj = wb_obj.active
                        max_col = sheet_obj.max_column
                        for i in range(1, max_col + 1):
                            cell_obj = sheet_obj.cell(row = 1, column = i)
                            if word in cell_obj.value:                            
                                self.find_file.append(x)
                                break          
                    except Exception as fail:
                        self.exc.append(r) 

        #pdf
        if format_files == 'pdf' or format_files == 'all': 
            for adress, dirs, files in os.walk(disk + ':\\'):
                for file in files:
                    s = (os.path.join(adress, file))
                    if file.endswith('.pdf') and '$' not in s:
                        self.spisok_file.append(s)
            for x in self.spisok_file: 
                with open(x) as r:
                    try:
                        warnings.filterwarnings('ignore', category=UserWarning, module='PyPDF2')
                        pdf_file = open(x, 'rb')
                        read_pdf = PyPDF2.PdfFileReader(pdf_file)
                        page = read_pdf.getPage(0)
                        page_content = page.extractText()
                        if word in page_content:                            
                            self.find_file.append(x)
                            break
                        pdf_file.close()
                    except Exception as fail:
                        self.exc.append(r) 

    #list of found files
    def list(self):
        if len(self.find_file) == 0:
            print('We did not find the words in files')
        else:
            return self.find_file

    #file opening
    def opening(self,opp = ''):    
        if len(self.find_file) == 0:
            print('We did not find the words in files')
        else:
            for x in self.find_file:
                head, tail = os.path.split(x)
                if opp == '':
                    print('We faind faile:', tail)
                    want_open = input('Open file? \n (y/n): ')
                    if want_open == 'y':
                        subprocess.Popen('explorer ' + x)
                elif opp != '':
                    subprocess.Popen('explorer ' + x)

    #copying found files
    def copying(self, path_copy = '', numcopy = 'a'):
        if len(self.find_file) == 0:
            print('We did not find the words in files')
        else:    
            try:
                os.mkdir(path_copy)
                for x in self.find_file:
                    head, tail = os.path.split(x) 
                    if numcopy != 'a':
                        print('Copy file', tail, '?' )
                        want_copy = input('y/n:')
                        if want_copy == 'y':
                            shutil.copy(x,path_copy)
                            print('File:', tail, 'was copied')
                    elif numcopy == 'a':
                        shutil.copy(x,path_copy)
                        print('File:', tail, 'was copied')
            except Exception as fail:
                print('A directory with this name already exists') 