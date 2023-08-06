import argparse
import csv
import os
from collections import namedtuple
from pathlib import Path
from typing import Optional
from typing import Sequence

import openpyxl
from slugify import slugify


OUT = Path(os.getenv("TMP", "/tmp"), "out")
WorkSheet = namedtuple("WorkSheet", "name rows")


def get_sheets(excel_file: Path) -> list[WorkSheet]:
    xlsx = openpyxl.load_workbook(excel_file)
    worksheets = []
    for name in xlsx.sheetnames:
        safe_name = slugify(name)
        rows = xlsx[name].iter_rows()
        worksheets.append(WorkSheet(name=safe_name, rows=rows))
    return worksheets


def sheets_to_csv(output_dir: Path, worksheets: list[WorkSheet]):
    for sheet in worksheets:
        output_dir.mkdir(parents=True, exist_ok=True)
        file = f"{output_dir}/{sheet.name}.csv"
        with open(file, "w") as fp:
            csvwriter = csv.writer(fp)
            for row in sheet.rows:
                csvwriter.writerow([cell.value for cell in row])


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "src",
        type=Path,
        nargs="?",
        help=("path to xlsx file."),
    )
    parser.add_argument(
        "dst",
        type=Path,
        default=OUT,
        nargs="?",
        help=(f"path to dest directory for csv files. (default: {OUT})"),
    )
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    excel_file = args.src
    output_dir = Path(args.dst / slugify(excel_file.name.partition(".")[0]))
    sheets = get_sheets(excel_file)
    sheets_to_csv(output_dir, sheets)
    return 0


if __name__ == "__main__":
    exit(main())
