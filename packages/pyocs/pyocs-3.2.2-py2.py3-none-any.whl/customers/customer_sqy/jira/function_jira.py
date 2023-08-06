#!/usr/bin/python
#coding:utf-8

#通过jira过滤器，捞出待审核的功能元，并添加到坚果云文档中

import os
import sys
PROJ_HOME_PATH = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(1,PROJ_HOME_PATH)

import re
import time
import json
import datetime
import requests
from pathlib import Path
from pprint import pprint

# import pandas as pd
# from pandas import DataFrame

import openpyxl

from pyocs import pyocs_jira
from pyocs.pyocs_software import PyocsSoftware
from pyocs.pyocs_confluence import PyocsConfluence
from pyocs.pyocs_filesystem import PyocsFileSystem
from pyocs.pyocs_login import PyocsLogin

CUR_PATH = os.path.dirname(os.path.abspath(__file__))
JiraSpecialMemberNameFile = CUR_PATH + "/jira_special_member_name.json"


#下面这些人的jira账号比较特殊
#需要做一下映射
with open(JiraSpecialMemberNameFile, 'r', encoding='UTF-8') as f:
    special_member = json.loads(f.read())

def download_file_form_drive(xlsx_file_path):
    #从坚果云下载:功能元.xls
    if Path(xlsx_file_path).exists() is True:
        os.remove(xlsx_file_path)
    map_list = PyocsConfluence().get_osm_distribute_mapping_info_by_user(xlsx_file_path.split('/')[-1])
    download_link = map_list[0]
    # print(download_link)
    # print(xlsx_file_path.split('/')[-1])
    PyocsFileSystem.get_file_from_nut_driver(download_link, '.')

def upload_to_nutstore(file_path: str):
    upload_nutstore_url = "https://drive.cvte.com/d/ajax/pubops/uploadXHR"

    form_data = dict()
    headers = dict()
    headers.update({
        'Origin': 'https://drive.cvte.com',
        'Referer': 'https://drive.cvte.com/p/DbXx7dEQz9ICGLCkCw',
        'Content-Type': 'application/octet-stream'
    })
    form_data.update({
        'key': 'DbXx7dEQz9ICGLCkCw',
        'dirName': '/',
        'path': '/' + file_path.split('/')[-1]
    })
    payload =  open(file_path, 'rb')
    r = requests.post(upload_nutstore_url, params=form_data,data=payload, headers=headers)
    return r.text

def write_to_xls_file(xls_file_name,content_dict_list):
    file_path = workspace + xls_file_name
    data = pd.read_excel(file_path)
    xls_len = len(data.values)

    for item in content_dict_list:
        xls_len = xls_len + 1

        title = item["title"]
        creator = item["creator"]
        create_time = item["create_time"]
        jira_link = item['jira_link']
        data.loc[xls_len] = [title, jira_link, creator, create_time, '','','']

    DataFrame(data).to_excel(workspace + xls_file_name, sheet_name='功能元', index=False, header=True)

def write_to_xlsx_file(xls_file_path,content_dict_list):
    wb=openpyxl.load_workbook(xls_file_path)
    shts=wb.get_sheet_names()
    sht=wb.get_sheet_by_name(shts[0])
    rows=sht.max_row
    columns=sht.max_column

    jira_list = list()
    for i in range(2,rows+1):
        if sht['B'+str(i)].value:
            actual_rows = i
            jira_list.append(sht['B'+str(i)].value)

    rows = actual_rows

    for item in content_dict_list:
        JIRA_AREADY_EXIST = False
        #检测xlsx中是否已包含待写入的功能元
        for j in jira_list:
            if item['jira_link'].split('/')[-1] in j:
                JIRA_AREADY_EXIST = True
                break
        if not JIRA_AREADY_EXIST:
            rows = rows + 1
            print("如下jira将被写入xlsx的第" + str(rows) + "行")
            print(item['jira_link'].split('/')[-1])
            sht['A'+str(rows)] = item["title"]
            sht['B'+str(rows)] = item['jira_link']
            sht['C'+str(rows)] = item["creator"]
            sht['D'+str(rows)] = item["create_time"]
            sht['E'+str(rows)] = ''
            sht['F'+str(rows)] = ''
            sht['G'+str(rows)] = ''

    wb.save(xls_file_path)


if __name__ == "__main__":

    xlsx_file_path = '功能元.xlsx'

    j = pyocs_jira.JiraCVTE()
    search_content = 'project = CUST_DOC_1 AND status in (新建,测试中, "In Progress") AND reporter not in (hanzhongyi, xionghui) ORDER BY priority DESC, updated DESC'
    jira_list = j.get_issue_jira_key_list(search_content)

    #从坚果云上下载xlsx文件
    download_file_form_drive(xlsx_file_path)

    #获取 待审核的功能元
    content_dict_list = []
    for issue_key in jira_list:
        issue_info = j.get_issue_jira_import_info(issue_key)

        content_dict = {}
        content_dict["title"] = issue_info["title"]
        content_dict["create_time"] = issue_info["create_time"]
        content_dict["jira_link"] = 'https://jira.cvte.com/projects/CUST_DOC_1/issues/' + issue_key
        if issue_info["creator"] in special_member.keys():
            issue_info["creator"] = special_member[issue_info["creator"]]
        content_dict["creator"] = PyocsSoftware().get_chinese_name_from_account(issue_info["creator"])
        content_dict_list.append(content_dict)

    write_to_xlsx_file(xlsx_file_path,content_dict_list)
    upload_to_nutstore(xlsx_file_path)




